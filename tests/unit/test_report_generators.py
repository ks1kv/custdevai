"""Smoke-тесты PDF/XLSX-генераторов на синтетическом контексте.

Не зависят от БД — собираем CampaignReportContext вручную и проверяем,
что генераторы возвращают валидные байтовые блобы (FR-RPT-01..05).
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

import pytest

from apps.api.db.models import SentimentLabel, SessionStatus
from apps.api.reports.data_loader import (
    AnswerView,
    CampaignReportContext,
    SessionView,
    TopicView,
)
from apps.api.reports.generators.pdf import PDFReportGenerator
from apps.api.reports.generators.xlsx import XLSXReportGenerator


def _make_context() -> CampaignReportContext:
    now = datetime(2026, 5, 8, 12, 0, tzinfo=timezone.utc)
    sessions = [
        SessionView(
            session_id=1,
            pseudonym="R-0001",
            status=SessionStatus.COMPLETED,
            started_at=now,
            completed_at=now,
        ),
        SessionView(
            session_id=2,
            pseudonym="R-0002",
            status=SessionStatus.COMPLETED,
            started_at=now,
            completed_at=now,
        ),
    ]
    answers = [
        AnswerView(
            answer_id=10,
            session_id=1,
            pseudonym="R-0001",
            question_id=100,
            question_text="Какие проблемы вы испытываете с продуктом?",
            question_order=1,
            text="Очень неудобный поиск, трачу полчаса вместо пяти минут.",
            answered_at=now,
            sentiment_label=SentimentLabel.NEGATIVE,
            sentiment_confidence=0.92,
        ),
        AnswerView(
            answer_id=11,
            session_id=2,
            pseudonym="R-0002",
            question_id=100,
            question_text="Какие проблемы вы испытываете с продуктом?",
            question_order=1,
            text="В целом всё устраивает, но интерфейс перегружен.",
            answered_at=now,
            sentiment_label=SentimentLabel.NEUTRAL,
            sentiment_confidence=0.77,
        ),
    ]
    topics = [
        TopicView(
            topic_id=1,
            topic_id_in_model=0,
            label="Юзабилити поиска",
            keywords=["поиск", "интерфейс", "удобство"],
            frequency_count=2,
            is_noise=False,
            quotes=[
                ("R-0001", "Очень неудобный поиск, трачу полчаса."),
                ("R-0002", "Интерфейс перегружен."),
            ],
        ),
    ]
    return CampaignReportContext(
        campaign_id=42,
        campaign_title="Исследование UX поиска",
        campaign_description="Глубинные интервью пользователей B2B-каталога.",
        script_title="Базовый сценарий CustDev",
        started_at=now,
        completed_at=now,
        target_topic_count=10,
        sessions=sessions,
        answers=answers,
        sentiment_distribution={SentimentLabel.NEGATIVE: 1, SentimentLabel.NEUTRAL: 1},
        topics=topics,
        generated_at=now,
    )


@pytest.fixture()
def context() -> CampaignReportContext:
    return _make_context()


class TestPDFReportGenerator:
    def test_returns_pdf_bytes(self, context: CampaignReportContext) -> None:
        # FR-RPT-02: PDF
        blob = PDFReportGenerator(context).render()
        assert isinstance(blob, bytes)
        assert blob[:4] == b"%PDF"
        assert len(blob) > 1000  # минимально содержательный PDF

    def test_pdf_has_minimum_structure(self, context: CampaignReportContext) -> None:
        # Базовая структура PDF: header + xref + trailer + EOF-маркер.
        # Текстовое содержимое компрессируется ReportLab — проверка цитат
        # с псевдонимами R-NNNN (FR-RPT-05) выполняется через XLSX-тест,
        # где значения ячеек доступны напрямую.
        blob = PDFReportGenerator(context).render()
        assert blob.startswith(b"%PDF-")
        assert b"%%EOF" in blob[-128:]
        assert b"xref" in blob


class TestXLSXReportGenerator:
    def test_returns_xlsx_bytes(self, context: CampaignReportContext) -> None:
        blob = XLSXReportGenerator(context).render()
        assert isinstance(blob, bytes)
        # XLSX — это zip-архив, начинается с PK\x03\x04.
        assert blob[:2] == b"PK"
        assert len(blob) > 1000

    def test_xlsx_has_three_sheets(self, context: CampaignReportContext) -> None:
        # FR-RPT-03: «Транскрипты», «Тональность», «Темы».
        from openpyxl import load_workbook

        blob = XLSXReportGenerator(context).render()
        wb = load_workbook(io.BytesIO(blob))
        assert set(wb.sheetnames) == {"Транскрипты", "Тональность", "Темы"}

    def test_xlsx_transcripts_contain_pseudonym(self, context: CampaignReportContext) -> None:
        from openpyxl import load_workbook

        blob = XLSXReportGenerator(context).render()
        wb = load_workbook(io.BytesIO(blob))
        ws = wb["Транскрипты"]
        col_a = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert "R-0001" in col_a
        assert "R-0002" in col_a
