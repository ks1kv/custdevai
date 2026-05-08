"""XLSXReportGenerator — выгрузка кампании в Excel с тремя листами (FR-RPT-03).

Листы (точно как в FR-RPT-03):
  1. «Транскрипты» — session R-NNNN | вопрос | ответ | answered_at |
     sentiment_label | confidence.
  2. «Тональность» — распределение по категориям + проценты.
  3. «Темы» — id | label | keywords | frequency | три цитаты.

NFR-OPS-06: все заголовки колонок и подписи на русском.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.api.db.models import SentimentLabel
from apps.api.reports.data_loader import CampaignReportContext

_HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
_HEADER_FONT = Font(bold=True)

_SENTIMENT_RU: dict[SentimentLabel, str] = {
    SentimentLabel.POSITIVE: "позитивная",
    SentimentLabel.NEUTRAL: "нейтральная",
    SentimentLabel.NEGATIVE: "негативная",
    SentimentLabel.LOW_CONFIDENCE: "низкая уверенность",
}


class XLSXReportGenerator:
    def __init__(self, ctx: CampaignReportContext) -> None:
        self._ctx = ctx

    def render(self) -> bytes:
        wb = Workbook()
        # Удаляем default-лист.
        default = wb.active
        if default is not None:
            wb.remove(default)
        self._sheet_transcripts(wb)
        self._sheet_sentiment(wb)
        self._sheet_topics(wb)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _sheet_transcripts(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Транскрипты")
        headers = [
            "Псевдоним",
            "№ вопроса",
            "Текст вопроса",
            "Ответ",
            "Время ответа",
            "Тональность",
            "Уверенность",
        ]
        ws.append(headers)
        self._format_header(ws, len(headers))

        for ans in self._ctx.answers:
            ws.append(
                [
                    ans.pseudonym,
                    ans.question_order + 1,
                    ans.question_text,
                    ans.text,
                    ans.answered_at.strftime("%Y-%m-%d %H:%M"),
                    _SENTIMENT_RU.get(ans.sentiment_label, "")
                    if ans.sentiment_label is not None
                    else "",
                    f"{ans.sentiment_confidence:.3f}"
                    if ans.sentiment_confidence is not None
                    else "",
                ]
            )
        # Авто-ширины колонок.
        widths = [12, 10, 40, 60, 18, 16, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        # Wrap для длинных текстов.
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _sheet_sentiment(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Тональность")
        ws.append(["Категория", "Число ответов", "Доля"])
        self._format_header(ws, 3)

        total = sum(self._ctx.sentiment_distribution.values())
        for label, count in self._ctx.sentiment_distribution.items():
            pct = (count / total) if total else 0
            row_idx = ws.max_row + 1
            ws.append([_SENTIMENT_RU.get(label, label.value), count, pct])
            ws.cell(row=row_idx, column=3).number_format = "0.0%"
        if not self._ctx.sentiment_distribution:
            ws.append(["Тональный анализ не проводился", "", ""])
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.freeze_panes = "A2"

    def _sheet_topics(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Темы")
        headers = [
            "ID темы",
            "Метка",
            "Ключевые слова",
            "Частота",
            "Шум",
            "Цитата 1 (псевдоним)",
            "Цитата 1 (текст)",
            "Цитата 2 (псевдоним)",
            "Цитата 2 (текст)",
            "Цитата 3 (псевдоним)",
            "Цитата 3 (текст)",
        ]
        ws.append(headers)
        self._format_header(ws, len(headers))

        for t in self._ctx.topics:
            row = [
                t.topic_id_in_model,
                t.label or "",
                ", ".join(t.keywords[:10]),
                t.frequency_count,
                "да" if t.is_noise else "нет",
            ]
            for i in range(3):
                if i < len(t.quotes):
                    row.extend(t.quotes[i])
                else:
                    row.extend(["", ""])
            ws.append(row)

        widths = [10, 24, 40, 12, 8, 14, 60, 14, 60, 14, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        for col_letter in ("G", "I", "K"):
            for cell in ws[col_letter]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _format_header(ws, num_cols: int) -> None:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
